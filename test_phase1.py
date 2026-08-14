from pathlib import Path

import pandas as pd

from app.engine import process_run
from app.output_writer import needs_secondary_file, write_xlsx, write_xlsx_secondary
from app.template_model import ColumnRule, Template

BASE = Path(__file__).parent
SAMPLE_DIR = BASE / "sample_data"
OUT_DIR = BASE / "test_output"

# --- synthetic source files (fake data, not real student records) ---
edmentum = pd.DataFrame(
    {
        "StudentID": ["1001", "1002", "1003", "1004"],
        "StudentName": ["Alice A", "Bob B", "Cara C", "Dan D"],
        "CourseName": ["Algebra I", "Biology", "US History", "English 9"],
        "PercentComplete": [82, 45, 100, 12],
    }
)
edmentum.to_csv(SAMPLE_DIR / "edmentum_export.csv", index=False)

# note: "Student ID" has a space (different header) and formatting (" 1002 ") to test
# normalized matching; row 1004 is missing here and 1005 has no Edmentum counterpart,
# both should surface as unmatched
ic = pd.DataFrame(
    {
        "Student ID": ["1001", " 1002 ", "1003", "1005"],
        "Student Name": ["Alice A", "Bob B", "Cara C", "Eve E"],
        "GradeLevel": [9, 10, 9, 11],
        "SchoolName": ["North HS", "North HS", "South HS", "South HS"],
    }
)
ic.to_excel(SAMPLE_DIR / "ic_export.xlsx", index=False)

# --- templates ---
edmentum_template = Template(
    name="Edmentum Course Export",
    columns=[
        ColumnRule("StudentID", delete=False, is_id=True),
        ColumnRule("StudentName", delete=True),
        ColumnRule("CourseName", delete=False),
        ColumnRule("PercentComplete", delete=False),
    ],
)

ic_template = Template(
    name="Infinite Campus Export",
    columns=[
        ColumnRule("Student ID", delete=False, is_id=True),
        ColumnRule("Student Name", delete=True),
        ColumnRule("GradeLevel", delete=False),
        ColumnRule("SchoolName", delete=False),
    ],
)

templates = {t.name: t for t in [edmentum_template, ic_template]}
input_files = {
    edmentum_template.name: SAMPLE_DIR / "edmentum_export.csv",
    ic_template.name: SAMPLE_DIR / "ic_export.xlsx",
}

result = process_run(templates, input_files, anonymize=True)

print("Unmatched rows by template:", result.unmatched_by_template)
print("Total unmatched:", result.unmatched_count)
for name, df in result.tables.items():
    print(f"\n--- {name} (matched) ---")
    print(df.to_string(index=False))
for name, df in result.unresolved_tables.items():
    print(f"\n--- Unresolved - {name} ---")
    print(df.to_string(index=False))

out_path = OUT_DIR / "anonymized_output.xlsx"
write_xlsx(result, out_path)
print(f"\nWrote: {out_path}")

secondary_path = OUT_DIR / "key_unresolved.xlsx"
assert needs_secondary_file(result, include_key=True), "this run has unresolved rows, should need a secondary file"
write_xlsx_secondary(result, secondary_path, include_key=True)
print(f"Wrote: {secondary_path}")

# template save/load round trip -- uses a throwaway store, never the app's real templates_store
import tempfile

with tempfile.TemporaryDirectory() as tmp_store_str:
    tmp_store = Path(tmp_store_str)
    edmentum_template.save(tmp_store)
    ic_template.save(tmp_store)
    loaded = Template.load_all(tmp_store)
    print("\nLoaded templates from disk:", [t.name for t in loaded])
    assert [t.name for t in loaded] == [t.name for t in [edmentum_template, ic_template]]
print("Template round-trip OK")

# sanity checks
assert result.unmatched_by_template["Edmentum Course Export"] == 1  # 1004 has no match
assert result.unmatched_by_template["Infinite Campus Export"] == 1  # 1005 has no match
assert result.unmatched_count == 2

# matched tabs should only contain the 3 matched rows, anonymized
assert len(result.tables["Edmentum Course Export"]) == 3
assert len(result.tables["Infinite Campus Export"]) == 3
row_e = result.tables["Edmentum Course Export"].iloc[0]  # was StudentID 1001
row_ic = result.tables["Infinite Campus Export"].iloc[0]  # was Student ID 1001
assert row_e["StudentID"] == row_ic["Student ID"], "Same real ID must map to same fake ID"
assert "StudentName" not in result.tables["Edmentum Course Export"].columns, "delete flag should drop column"

# unresolved tabs should hold exactly the unmatched row, unedited (raw ID, deleted columns intact)
assert list(result.unresolved_tables.keys()) == ["Edmentum Course Export", "Infinite Campus Export"]
unresolved_e = result.unresolved_tables["Edmentum Course Export"]
assert len(unresolved_e) == 1
assert unresolved_e.iloc[0]["StudentID"] == "1004", "unresolved rows must keep the real, unanonymized ID"
assert "StudentName" in unresolved_e.columns, "unresolved rows must be unedited -- no column deletion"
unresolved_ic = result.unresolved_tables["Infinite Campus Export"]
assert unresolved_ic.iloc[0]["Student ID"] == "1005"

# key table should map only the 3 matched real IDs to their anonymized IDs -- not the
# 2 unresolved ones, since those never got anonymized in the output
assert sorted(result.key_table["Common Identifier"]) == ["1001", "1002", "1003"]
key_lookup = dict(zip(result.key_table["Common Identifier"], result.key_table["Anonymized ID"]))
assert key_lookup["1001"] == row_e["StudentID"], "key must map the real ID to the same fake ID used in the output"

# primary (clean) output should have ONLY the 2 matched sheets -- never anything
# unresolved or identifying, so it's always safe to hand off on its own
import openpyxl  # noqa: E402

wb = openpyxl.load_workbook(out_path)
assert set(wb.sheetnames) == {"Edmentum Course Export", "Infinite Campus Export"}, wb.sheetnames

# secondary file should hold the unresolved tabs + the key (Excel sheet names cap at 31 chars)
wb_secondary = openpyxl.load_workbook(secondary_path)
assert set(wb_secondary.sheetnames) == {
    "Unresolved - Edmentum Course Ex",
    "Unresolved - Infinite Campus Ex",
    "Key",
}, wb_secondary.sheetnames
key_ws = wb_secondary["Key"]
key_rows = list(key_ws.iter_rows(values_only=True))
assert key_rows[0] == ("Common Identifier", "Anonymized ID")
assert len(key_rows) == 4  # header + 3 matched IDs

# a run with nothing unresolved and no key requested shouldn't need a secondary file
no_secondary_result = process_run(
    {edmentum_template.name: edmentum_template},
    {edmentum_template.name: SAMPLE_DIR / "edmentum_export.csv"},
    anonymize=False,
)
assert not needs_secondary_file(no_secondary_result, include_key=False)

# --- regression test: a blank common identifier must go to Unresolved, never silently
# ride into the matched/anonymized tab with an empty ID ---
with tempfile.TemporaryDirectory() as tmp_dir_str:
    tmp_dir = Path(tmp_dir_str)
    blank_a = pd.DataFrame({"ID": ["A1", "", "A3"], "Val": ["x", "y", "z"]})
    blank_a.to_csv(tmp_dir / "blank_a.csv", index=False)
    blank_b = pd.DataFrame({"ID": ["A1", "A2", "A3"], "Val": ["p", "q", "r"]})
    blank_b.to_csv(tmp_dir / "blank_b.csv", index=False)

    t_a = Template(name="Blank A", columns=[ColumnRule("ID", delete=False, is_id=True), ColumnRule("Val", delete=False)])
    t_b = Template(name="Blank B", columns=[ColumnRule("ID", delete=False, is_id=True), ColumnRule("Val", delete=False)])

    blank_result = process_run(
        {t_a.name: t_a, t_b.name: t_b},
        {t_a.name: tmp_dir / "blank_a.csv", t_b.name: tmp_dir / "blank_b.csv"},
        anonymize=True,
    )
    assert len(blank_result.tables["Blank A"]) == 2, "blank-ID row must not appear in the matched tab"
    assert "Blank A" in blank_result.unresolved_tables
    assert pd.isna(blank_result.unresolved_tables["Blank A"].iloc[0]["ID"]), "unresolved row must keep the blank ID"
print("Blank-ID regression check passed.")

print("\nAll Phase 1 sanity checks passed.")
